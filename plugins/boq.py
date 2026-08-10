"""P02 — Bill-of-Quantities intake (real implementation).

This capability turns an ingested spreadsheet source into structured
bill-of-quantities (BOQ) line items, each expressed as the five elements
required by GB 50500《建设工程工程量清单计价规范》:

    编码 code · 名称 name · 项目特征 feature · 计量单位 unit · 工程量 quantity

Two input shapes are supported through an explicit ``mode``:

* ``"standard"`` (default): the source is already a formed GB 50500 BOQ
  table; rows are read and the 9-digit item codes are validated.
* ``"merge"``: the source is a 广联达-style "绘图输入工程量汇总表" where
  quantities are scattered across per-component sheets; a caller-supplied
  mapping table (plain JSON, never pickle) merges component rows into BOQ
  items. This logic is distilled from the 诚清单 tool's matching-and-merge
  behaviour and rewritten natively — no xlwings, no pickle, no Qt.

The capability performs pure, side-effect-free parsing. It reads bytes that
Core has already content-addressed and returns plain result rows; it never
writes files and never executes external processes.

Design notes (why this differs from the distilled source):
* xlwings → openpyxl: no Excel process, runs headless in any container.
* pickle mapping → JSON mapping: loading a pickle is arbitrary code
  execution; JSON is inert data.
* Qt GUI → removed entirely: BuildCostIQ's gateway only accepts bytes in
  and structured rows out; presentation is out of scope for a capability.
"""

from __future__ import annotations

import io
import json
import re
from collections.abc import Mapping, Sequence
from typing import Any

try:  # openpyxl is the only new runtime dependency, and only for .xlsx input.
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - surfaced as a clear runtime error.
    load_workbook = None


# GB 50500 clause codes are 9 digits (国标 12 位在清单中常截取前 9 位作为项目编码前缀).
# We accept 9-to-12 digit numeric codes and treat the leading 9 as the classifier.
_CODE_PATTERN = re.compile(r"^\d{9,12}$")

# Canonical column keys for a standard BOQ row, mapped from common header aliases.
_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "code": ("项目编码", "清单编码", "编码", "code"),
    "name": ("项目名称", "清单名称", "名称", "name"),
    "feature": ("项目特征", "特征描述", "特征", "feature"),
    "unit": ("计量单位", "单位", "unit"),
    "quantity": ("工程量", "工程数量", "数量", "quantity"),
}


class BoqParseError(ValueError):
    """Raised when a source cannot be parsed into valid BOQ items."""


def _require_openpyxl() -> None:
    if load_workbook is None:
        raise BoqParseError(
            "openpyxl is required to parse .xlsx sources; install it or pass pre-parsed rows"
        )


def _normalize_header(cell: object) -> str:
    return str(cell).strip() if cell is not None else ""


def _resolve_columns(header_row: Sequence[object]) -> dict[str, int]:
    """Map canonical field -> column index using header aliases."""
    normalized = [_normalize_header(c) for c in header_row]
    resolved: dict[str, int] = {}
    for field, aliases in _HEADER_ALIASES.items():
        for idx, head in enumerate(normalized):
            if head in aliases:
                resolved[field] = idx
                break
    return resolved


def _clean_quantity(value: object) -> float:
    """Convert a cell value to a non-negative float quantity."""
    if value is None or value == "":
        raise BoqParseError("empty quantity")
    try:
        qty = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError) as exc:
        raise BoqParseError(f"non-numeric quantity: {value!r}") from exc
    if qty < 0:
        raise BoqParseError(f"negative quantity: {qty}")
    return qty


def _validate_code(code: object) -> str:
    text = str(code).strip()
    if not _CODE_PATTERN.match(text):
        raise BoqParseError(f"invalid GB 50500 item code: {code!r}")
    return text


def parse_standard_boq(rows: Sequence[Sequence[object]]) -> list[dict[str, Any]]:
    """Parse a formed GB 50500 BOQ table (header row + data rows).

    Returns one dict per line item with the five canonical elements.
    Rows whose code cell is blank are treated as section headers and skipped.
    """
    if not rows:
        return []
    columns = _resolve_columns(rows[0])
    missing = [f for f in ("code", "name", "unit", "quantity") if f not in columns]
    if missing:
        raise BoqParseError(f"BOQ header missing required columns: {', '.join(missing)}")

    items: list[dict[str, Any]] = []
    for line_no, row in enumerate(rows[1:], start=2):
        raw_code = row[columns["code"]] if columns["code"] < len(row) else None
        # A priced line item has a valid GB 50500 numeric code. Anything else
        # in the code column (blank, a section title like "措施项目", a
        # subtotal label) is structural and skipped rather than rejected.
        if raw_code is None or not _CODE_PATTERN.match(str(raw_code).strip()):
            continue
        try:
            item = {
                "code": _validate_code(raw_code),
                "name": str(row[columns["name"]]).strip(),
                "feature": (
                    str(row[columns["feature"]]).strip()
                    if "feature" in columns and columns["feature"] < len(row) and row[columns["feature"]] is not None
                    else ""
                ),
                "unit": str(row[columns["unit"]]).strip(),
                "quantity": _clean_quantity(row[columns["quantity"]]),
            }
        except BoqParseError as exc:
            raise BoqParseError(f"row {line_no}: {exc}") from exc
        items.append(item)
    return items


def parse_merge_boq(
    sheets: Mapping[str, Sequence[Sequence[object]]],
    mapping: Mapping[str, Mapping[str, str]],
) -> list[dict[str, Any]]:
    """Merge 广联达-style per-component sheets into BOQ items via a JSON mapping.

    ``mapping`` shape (plain JSON, no pickle):

        {
          "<sheet name>": {
            "code": "010502001001",
            "name": "矩形柱",
            "feature": "C30 现浇",
            "unit": "m3",
            "quantity_keyword": "体积"     # header cell to sum within the sheet
          },
          ...
        }

    For each mapped sheet, the column under ``quantity_keyword`` is summed
    across data rows to produce the BOQ item's quantity. This is the distilled
    core of 诚清单's matching table, expressed declaratively.
    """
    items: list[dict[str, Any]] = []
    for sheet_name, rule in mapping.items():
        if sheet_name not in sheets:
            continue
        table = sheets[sheet_name]
        if not table:
            continue
        header = [_normalize_header(c) for c in table[0]]
        keyword = rule.get("quantity_keyword", "工程量")
        try:
            qty_col = header.index(keyword)
        except ValueError as exc:
            raise BoqParseError(
                f"sheet {sheet_name!r}: quantity column {keyword!r} not found"
            ) from exc

        total = 0.0
        for row in table[1:]:
            if qty_col >= len(row):
                continue
            # Skip subtotal/total rows: 广联达 sheets end each section with a
            # 小计/合计 row whose quantity equals the section sum, so counting
            # it would double the total. Detect the label anywhere in the row.
            if any(
                isinstance(c, str) and any(tag in c for tag in ("小计", "合计", "总计"))
                for c in row
            ):
                continue
            cell = row[qty_col]
            if cell is None or str(cell).strip() == "":
                continue
            try:
                total += _clean_quantity(cell)
            except BoqParseError:
                continue  # skip non-numeric rows (labels, blanks) inside the sheet

        items.append(
            {
                "code": _validate_code(rule["code"]),
                "name": str(rule.get("name", sheet_name)).strip(),
                "feature": str(rule.get("feature", "")).strip(),
                "unit": str(rule.get("unit", "")).strip(),
                "quantity": round(total, 6),
            }
        )
    return items


def _read_xlsx(content: bytes) -> dict[str, list[list[object]]]:
    """Read all sheets of an .xlsx byte stream into plain nested lists."""
    _require_openpyxl()
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        return {
            sheet.title: [list(row) for row in sheet.iter_rows(values_only=True)]
            for sheet in workbook.worksheets
        }
    finally:
        workbook.close()


class BillOfQuantitiesCapability:
    """P02 — Bill-of-quantities intake.

    Context keys:
        project_id (str, required)
        source_id  (str, required)
        boq_bytes  (bytes, required for xlsx input) — the ingested .xlsx source
        mode       (str, optional) — "standard" (default) or "merge"
        mapping    (dict, required when mode == "merge") — JSON mapping table
        rows       (list, optional) — pre-parsed rows, bypasses xlsx reading
                                      (used for standard mode without openpyxl)

    Returns a mapping with the capability id, accepted status, item count,
    and the parsed BOQ items ready to be recorded as Evidence(kind="boq_item").
    """

    capability_id = "P02"
    name = "bill-of-quantities"

    def execute(self, context: Mapping[str, Any]) -> Mapping[str, Any]:
        required = ("project_id", "source_id")
        missing = [key for key in required if not context.get(key)]
        if missing:
            raise ValueError(f"Missing context: {', '.join(missing)}")

        mode = context.get("mode", "standard")

        if mode == "standard":
            rows = context.get("rows")
            content = context.get("boq_bytes")
            if rows is None and isinstance(content, (bytes, bytearray)):
                sheets = _read_xlsx(bytes(content))
                # Use the first non-empty sheet as the BOQ table.
                rows = next((t for t in sheets.values() if t), [])
            # No source yet is a valid "nothing to intake" state, not an error;
            # this keeps the capability traversable with a bare context.
            items = parse_standard_boq(rows) if rows else []

        elif mode == "merge":
            mapping = context.get("mapping")
            if isinstance(mapping, (bytes, bytearray, str)):
                mapping = json.loads(mapping)
            if not isinstance(mapping, Mapping):
                raise ValueError("P02 merge mode needs a JSON 'mapping' object")
            content = context.get("boq_bytes")
            if not isinstance(content, (bytes, bytearray)):
                raise ValueError("P02 merge mode needs 'boq_bytes'")
            sheets = _read_xlsx(bytes(content))
            items = parse_merge_boq(sheets, mapping)

        else:
            raise ValueError(f"Unknown P02 mode: {mode!r}")

        return {
            "capability_id": self.capability_id,
            "name": self.name,
            "project_id": context["project_id"],
            "source_id": context["source_id"],
            "status": "accepted",
            "item_count": len(items),
            "items": items,
        }
