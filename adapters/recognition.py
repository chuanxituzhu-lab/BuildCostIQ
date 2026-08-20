"""Local-first source recognition and explicit external OCR adapters.

The source store remains the system of record. Recognition creates a local
derived Markdown artifact and metadata; it never replaces the original file.
External providers are opt-in per request and never run from file upload.
"""

from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from io import BytesIO
import base64
import html
from importlib.util import find_spec
import json
import os
from pathlib import Path
import re
from typing import Any
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from zipfile import BadZipFile, ZipFile
from xml.etree import ElementTree


IMAGE_SUFFIXES = {".bmp", ".gif", ".jpeg", ".jpg", ".png", ".tif", ".tiff", ".webp"}
MARKITDOWN_SUFFIXES = {".csv", ".doc", ".docx", ".epub", ".html", ".json", ".pdf", ".pptx", ".txt", ".xlsx", ".xml", ".zip"}


class RecognitionError(ValueError):
    """A source cannot be recognized by the selected local adapter."""


@dataclass(frozen=True, slots=True)
class RecognitionDescriptor:
    id: str
    name: str
    mode: str
    directions: tuple[str, ...]
    formats: tuple[str, ...]
    status: str
    requires_explicit_consent: bool
    description: str


RECOGNIZERS: tuple[RecognitionDescriptor, ...] = (
    RecognitionDescriptor(
        "local-auto",
        "本地自动识别",
        "local",
        ("recognize",),
        (".pdf", ".docx", ".xlsx", ".csv", ".dxf", ".txt", ".md", ".html", ".jpg", ".png"),
        "ready",
        False,
        "先在本机提取文字、表格、图纸文本和文件元数据，并自动归档分类。",
    ),
    RecognitionDescriptor(
        "microsoft-markitdown",
        "Microsoft MarkItDown（本地）",
        "local",
        ("recognize",),
        tuple(sorted(MARKITDOWN_SUFFIXES)),
        "optional",
        False,
        "在本机把 PDF、Office、HTML、CSV 等资料转换为 Markdown，作为可检索副本。",
    ),
    RecognitionDescriptor(
        "baidu-ocr",
        "百度 OCR",
        "external",
        ("recognize",),
        tuple(sorted(IMAGE_SUFFIXES)),
        "configured" if os.environ.get("BUILDCOSTIQ_BAIDU_API_KEY") and os.environ.get("BUILDCOSTIQ_BAIDU_SECRET_KEY") else "requires_configuration",
        True,
        "仅在用户明确确认后，把指定图片发送到百度 OCR；默认不发送。",
    ),
)


def recognition_catalog() -> list[dict[str, Any]]:
    markitdown_available = _markitdown_is_available()
    baidu_configured = bool(os.environ.get("BUILDCOSTIQ_BAIDU_API_KEY") and os.environ.get("BUILDCOSTIQ_BAIDU_SECRET_KEY"))
    catalog: list[dict[str, Any]] = []
    for item in RECOGNIZERS:
        row = asdict(item)
        if item.id == "microsoft-markitdown":
            row["status"] = "ready" if markitdown_available else "optional"
        elif item.id == "baidu-ocr":
            row["status"] = "configured" if baidu_configured else "requires_configuration"
        catalog.append(row)
    return catalog


def _markitdown_is_available() -> bool:
    # Catalog reads must stay fast; importing MarkItDown can load optional
    # Office/PDF dependencies. The actual converter path still imports it.
    return find_spec("markitdown") is not None


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _pdf_text(content: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - depends on local install
        raise RecognitionError("本地 PDF 识别需要 pypdf") from exc
    try:
        reader = PdfReader(BytesIO(content), strict=False)
        pages = [page.extract_text() or "" for page in reader.pages]
    except Exception as exc:  # pragma: no cover - malformed third-party PDFs
        raise RecognitionError(f"PDF 无法读取：{exc}") from exc
    return "\n\n".join(page.strip() for page in pages if page.strip())


def _docx_text(content: bytes) -> str:
    try:
        with ZipFile(BytesIO(content)) as archive:
            xml = archive.read("word/document.xml")
    except (BadZipFile, KeyError, OSError) as exc:
        raise RecognitionError("Word 文件无法读取") from exc
    try:
        root = ElementTree.fromstring(xml)
    except ElementTree.ParseError as exc:
        raise RecognitionError("Word 文档结构无法读取") from exc
    paragraphs: list[str] = []
    for paragraph in root.iter():
        if not paragraph.tag.endswith("}p"):
            continue
        text = "".join(node.text or "" for node in paragraph.iter() if node.tag.endswith("}t"))
        if text.strip():
            paragraphs.append(text.strip())
    return "\n\n".join(paragraphs)


def _xlsx_text(content: bytes) -> str:
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - malformed third-party workbooks
        raise RecognitionError(f"Excel 文件无法读取：{exc}") from exc
    sections: list[str] = []
    for sheet in workbook.worksheets:
        sections.append(f"## {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if values:
                sections.append(" | ".join(values))
    return "\n".join(sections)


def _dxf_text(content: bytes) -> str:
    text = _decode_text(content)
    lines = [line.strip() for line in text.splitlines()]
    layers: set[str] = set()
    entities: set[str] = set()
    for index, code in enumerate(lines[:-1]):
        value = lines[index + 1]
        if code == "8" and value:
            layers.add(value)
        if code == "0" and value:
            entities.add(value)
    details = ["# CAD 图纸识别结果", f"- 图层数：{len(layers)}", f"- 实体类型数：{len(entities)}"]
    if layers:
        details.append(f"- 图层：{', '.join(sorted(layers)[:80])}")
    if entities:
        details.append(f"- 实体：{', '.join(sorted(entities)[:80])}")
    return "\n".join(details)


def _markitdown_text(name: str, content: bytes) -> str:
    try:
        from markitdown import MarkItDown, StreamInfo
    except ImportError as exc:  # pragma: no cover - optional local install
        raise RecognitionError("本地 MarkItDown 未安装") from exc
    try:
        converter = MarkItDown(enable_plugins=False)
        result = converter.convert_stream(
            BytesIO(content),
            stream_info=StreamInfo(extension=Path(name).suffix, filename=Path(name).name),
        )
    except Exception as exc:  # pragma: no cover - optional converter/runtime
        raise RecognitionError(f"MarkItDown 转换失败：{exc}") from exc
    text = getattr(result, "text_content", None) or getattr(result, "markdown", None) or ""
    return str(text).strip()


def _local_text(name: str, content: bytes, connector_id: str) -> tuple[str, str]:
    suffix = Path(name).suffix.lower()
    simple_local_suffixes = {".txt", ".md", ".csv", ".html", ".htm", ".json", ".xml", ".log"}
    if connector_id == "microsoft-markitdown" or (connector_id == "local-auto" and suffix in MARKITDOWN_SUFFIXES and suffix not in {".pdf", ".docx", ".xlsx", *simple_local_suffixes}):
        try:
            text = _markitdown_text(name, content)
            if text:
                return text, "microsoft-markitdown"
        except RecognitionError:
            if connector_id == "microsoft-markitdown":
                raise
    if suffix == ".pdf":
        return _pdf_text(content), "pypdf-local"
    if suffix == ".docx":
        return _docx_text(content), "docx-xml-local"
    if suffix in {".xlsx", ".xlsm"}:
        return _xlsx_text(content), "openpyxl-local"
    if suffix == ".dxf":
        return _dxf_text(content), "dxf-text-local"
    if suffix in IMAGE_SUFFIXES:
        return "", "local-image-metadata-only"
    if suffix in {".txt", ".md", ".csv", ".html", ".htm", ".json", ".xml", ".log"}:
        return _decode_text(content), "text-local"
    raise RecognitionError(f"暂不支持本地识别：{suffix or '未知格式'}")


def _classify(name: str, text: str) -> tuple[str, list[str], float]:
    haystack = f"{name}\n{text[:120_000]}".lower()
    groups = {
        "清单与计价": ("清单", "工程量", "项目编码", "单价", "金额", "boq", "quantity", "price"),
        "合同与商务": ("合同", "协议", "中标", "投标", "结算", "付款", "contract", "tender"),
        "图纸与算量": ("图纸", "算量", "cad", "dwg", "dxf", "平面", "立面", "section"),
        "规范与文章": ("规范", "标准", "规程", "办法", "通知", "文章", "论文", "specification"),
        "现场与影像": ("现场", "照片", "影像", "签证", "变更", "image", "photo"),
        "结算与审计": ("结算", "审计", "审价", "核对", "初审", "settlement", "audit"),
    }
    scores = {category: sum(haystack.count(keyword.lower()) for keyword in keywords) for category, keywords in groups.items()}
    category = max(scores, key=scores.get) if any(scores.values()) else "项目资料"
    score = scores.get(category, 0)
    confidence = min(0.98, 0.45 + score * 0.08) if score else 0.35
    tags = [keyword for keyword in groups.get(category, ()) if keyword.lower() in haystack][:8]
    suffix = Path(name).suffix.lower()
    if suffix in {".pdf", ".doc", ".docx", ".txt", ".md", ".html"} and category == "项目资料":
        category = "文章与文档"
    if suffix in IMAGE_SUFFIXES and category == "项目资料":
        category = "现场与影像"
    if suffix in {".dwg", ".dxf"}:
        category = "图纸与算量"
    return category, tags, confidence


def _markdown(name: str, category: str, text: str) -> str:
    title = Path(name).stem or name
    return f"# {title}\n\n- 来源文件：{name}\n- 自动归档：{category}\n\n{text.strip()}\n"


def _local_result(name: str, content: bytes, connector_id: str) -> tuple[dict[str, Any], bytes | None]:
    suffix = Path(name).suffix.lower()
    try:
        text, engine = _local_text(name, content, connector_id)
    except RecognitionError as exc:
        status = "needs_ocr" if suffix in IMAGE_SUFFIXES or suffix == ".pdf" else "unavailable"
        result = {
            "status": status,
            "mode": "local",
            "connector_id": connector_id,
            "engine": "local-first",
            "category": "现场与影像" if suffix in IMAGE_SUFFIXES else "项目资料",
            "tags": [],
            "confidence": 0.2,
            "text_length": 0,
            "text_preview": "",
            "message": str(exc),
            "recognized_at": _now(),
        }
        return result, None
    if not text.strip():
        result = {
            "status": "needs_ocr" if suffix in IMAGE_SUFFIXES or suffix == ".pdf" else "completed",
            "mode": "local",
            "connector_id": connector_id,
            "engine": engine,
            "category": "现场与影像" if suffix in IMAGE_SUFFIXES else "项目资料",
            "tags": [],
            "confidence": 0.25,
            "text_length": 0,
            "text_preview": "",
            "message": "本地没有提取到文字，可在明确授权后选择外部 OCR。",
            "recognized_at": _now(),
        }
        return result, None
    category, tags, confidence = _classify(name, text)
    markdown = _markdown(name, category, text)
    result = {
        "status": "completed",
        "mode": "local",
        "connector_id": connector_id,
        "engine": engine,
        "category": category,
        "tags": tags,
        "confidence": confidence,
        "text_length": len(text),
        "text_preview": text[:500],
        "message": "已在本地提取文字并建立 Markdown 归档副本。",
        "recognized_at": _now(),
    }
    return result, markdown.encode("utf-8")


def _baidu_ocr(name: str, content: bytes) -> tuple[dict[str, Any], bytes | None]:
    suffix = Path(name).suffix.lower()
    if suffix not in IMAGE_SUFFIXES:
        raise RecognitionError("百度 OCR 适配器当前接收图片文件；扫描 PDF 请先选择本地转换或专用文档 OCR。")
    api_key = os.environ.get("BUILDCOSTIQ_BAIDU_API_KEY", "")
    secret_key = os.environ.get("BUILDCOSTIQ_BAIDU_SECRET_KEY", "")
    if not api_key or not secret_key:
        return {
            "status": "not_configured",
            "mode": "external",
            "connector_id": "baidu-ocr",
            "category": "现场与影像",
            "tags": [],
            "confidence": 0.0,
            "text_length": 0,
            "text_preview": "",
            "message": "未配置百度 OCR 凭据；原文件没有发送。",
            "recognized_at": _now(),
        }, None
    token_query = urlencode({"grant_type": "client_credentials", "client_id": api_key, "client_secret": secret_key})
    with urlopen(f"https://aip.baidubce.com/oauth/2.0/token?{token_query}", timeout=20) as response:
        token_payload = json.loads(response.read().decode("utf-8"))
    access_token = token_payload.get("access_token")
    if not access_token:
        raise RecognitionError("百度 OCR access token 获取失败")
    payload = urlencode({"image": base64.b64encode(content).decode("ascii")}).encode("ascii")
    request = Request(
        f"https://aip.baidubce.com/rest/2.0/ocr/v1/general_basic?access_token={access_token}",
        data=payload,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    with urlopen(request, timeout=30) as response:
        ocr_payload = json.loads(response.read().decode("utf-8"))
    text = "\n".join(str(item.get("words", "")) for item in ocr_payload.get("words_result", []) if item.get("words"))
    category, tags, confidence = _classify(name, text)
    result = {
        "status": "completed" if text else "needs_ocr",
        "mode": "external",
        "connector_id": "baidu-ocr",
        "engine": "baidu-general-basic",
        "category": category,
        "tags": tags,
        "confidence": confidence if text else 0.2,
        "text_length": len(text),
        "text_preview": text[:500],
        "message": "已按用户授权调用百度 OCR；原文件仍保存在本地。" if text else "百度 OCR 未返回文字。",
        "recognized_at": _now(),
    }
    return result, _markdown(name, category, text).encode("utf-8") if text else None


def recognize_source(name: str, content: bytes, connector_id: str = "local-auto", allow_external: bool = False) -> tuple[dict[str, Any], bytes | None]:
    """Recognize one source; external adapters are blocked without explicit consent."""
    if connector_id == "baidu-ocr":
        if not allow_external:
            return {
                "status": "consent_required",
                "mode": "external",
                "connector_id": connector_id,
                "category": "现场与影像",
                "tags": [],
                "confidence": 0.0,
                "text_length": 0,
                "text_preview": "",
                "message": "此操作会把指定文件发送到百度 OCR，请明确确认后再执行。",
                "recognized_at": _now(),
            }, None
        return _baidu_ocr(name, content)
    if connector_id not in {"local-auto", "microsoft-markitdown"}:
        raise RecognitionError(f"未知识别工具：{connector_id}")
    return _local_result(name, content, connector_id)
